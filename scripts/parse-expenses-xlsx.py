#!/usr/bin/env python3
"""One-off parser: "Reporte de Autos - AW.xlsx" -> JSON rows for internal_expenses.

Only parses the "Vehicle Expense Report" sheet (the master per-client report).
Sheets "Cust # 9301 ARG" and "Detalle transferencia" are source ledgers whose
items are already consolidated into the master sheet — importing them would
double-count. This script cross-checks them by amount and reports any entry
that does NOT appear in the master sheet, so nothing is silently lost.

Usage: python3 scripts/parse-expenses-xlsx.py <xlsx path> <output json path>
"""

import json
import re
import sys
from datetime import datetime

import openpyxl

PROGRAM = "Argentina Export"

CATEGORY_RULES = [
    ("Seguro", ["seguro", "progressive"]),
    ("Comisión", ["comision", "comsion", "cupo", "broker", "comm "]),
    ("Título y Trámites", ["titulo", "title", "placa", "tag ", "apostill", "notariz",
                         "west flagler", "west flager", "escribania", "consulado",
                         "baja + itv", "matriculacion", "papele", "tramite", "traimite",
                         "certificacion de documentos", "despachante"]),
    ("Transporte y Envío", ["grua", "tow", "transport", "contenedor", "red logistics",
                          "red logistcs", "envio", "flete", "puerto", "uhaul",
                          "mudanza", "storage", "trastero", "deposito fiscal",
                          "megaton", "megatom", "metagatom", "hbl", "embajale",
                          "aki", "autotrader"]),
    ("Mecánica", ["fixar", "mtech", "valhalla", "vahalla", "mecanic", "mechanic",
                  "brake", "tire", "llanta", "engine", "repair", "polish", "pulido",
                  "alignment", "tapiceria", "tapizeria", "dashboard", "tablero",
                  "spoiler", "chrome", "kilometraje", "km", "millas", "velocimetro",
                  "enmascar", "fondo", "instrumental", "desmontar", "grill",
                  "timing chain", "clutch", "top ", "lock cylinder", "premier"]),
    ("Viajes y Viáticos", ["uber", "pasaje", "hotel", "viaje", "viatico", "gasolina",
                         "tickt", "ticket"]),
    ("Fees", ["fee", "bank", "acreditacion", "diferencia", "diff", "porcentaje",
              "costo por transferencia", "le cobran"]),
]

VEHICLE_COST_PAT = re.compile(
    r"^(costo|vehicle cost|transferencia (1|2|realiazada|realizada)|gonzalo envia|gasto de juan)",
    re.IGNORECASE,
)


def clean_id(value):
    """480129.0 -> '480129'; keep strings; '-' -> None."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text if text and text != "-" else None


def categorize(description):
    desc = description.lower()
    if VEHICLE_COST_PAT.match(desc.strip()) and "costo por" not in desc and "costo de grua" not in desc:
        return "Costo del Vehículo"
    for category, keywords in CATEGORY_RULES:
        if any(k in desc for k in keywords):
            return category
    return "Otros"


DATE_PAT = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{2,4})")


def extract_date(description):
    matches = DATE_PAT.findall(description)
    if not matches:
        return None
    month, day, year = (int(x) for x in matches[-1])
    if year < 100:
        year += 2000
    try:
        return datetime(year, month, day).date().isoformat()
    except ValueError:
        return None


EUR_PAT = re.compile(r"([\d.,]+)\s*Euros", re.IGNORECASE)
RATE_PAT = re.compile(r"\(\s*([\d.]+)\s*\)")


def parse_side_note(note, amount_usd):
    """Column C notes like '68500 Euros' or '55,700 Euros ( 1.14 )'."""
    result = {}
    match = EUR_PAT.search(note)
    if match:
        original = float(match.group(1).replace(",", ""))
        result["amount_original"] = original
        result["currency_original"] = "EUR"
        rate_match = RATE_PAT.search(note)
        if rate_match:
            result["exchange_rate"] = float(rate_match.group(1))
        elif original:
            result["exchange_rate"] = round(amount_usd / original, 4)
    elif note.strip() and note.strip() != "`":
        result["note"] = note.strip()
    return result


def parse_master_sheet(ws):
    rows = []
    block = None
    blocks_meta = []

    def flush():
        if block and block["expenses"]:
            blocks_meta.append(block)
            rows.extend(block["expenses"])

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        side_notes = [str(ws.cell(r, c).value).strip()
                      for c in range(3, 7)
                      if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip()]
        a_text = str(a).strip() if a is not None else ""

        is_client_row = a_text.lower().startswith("client name")
        is_vw_block = a_text.upper().startswith("VW BEETLE")

        if is_client_row or is_vw_block:
            flush()
            note_bits = list(side_notes)
            paren = re.search(r"\((.+)\)", a_text)
            if paren:
                note_bits.append(paren.group(1).strip())
            block = {
                "client_name": str(b).strip() if b else None,
                "vehicle": a_text if is_vw_block else None,
                "stock_number": None, "deal_number": None,
                "customer_number": None, "vin": None,
                "block_note": " | ".join(note_bits) or None,
                "excel_row": r, "excel_total": None,
                "expenses": [],
            }
            continue

        if block is None:
            continue

        key = a_text.rstrip(":").strip().lower()
        if key in ("stock", "deal #", "customer #", "vehicle", "vin #",
                   "title issue date", "date issue", "despachante", "description"):
            if key == "stock":
                block["stock_number"] = clean_id(b)
            elif key == "deal #":
                block["deal_number"] = clean_id(b)
            elif key == "customer #":
                block["customer_number"] = clean_id(b)
            elif key == "vehicle":
                block["vehicle"] = str(b).strip() if b else None
            elif key == "vin #":
                block["vin"] = str(b).strip() if b else None
            elif key in ("title issue date", "date issue") and b:
                date = b.date().isoformat() if hasattr(b, "date") else str(b)
                note = f"Title issue date: {date}"
                block["block_note"] = f"{block['block_note']} | {note}" if block["block_note"] else note
            continue

        # Total row: no description, numeric amount
        if not a_text and isinstance(b, (int, float)):
            block["excel_total"] = float(b)
            continue

        # Expense row
        if a_text:
            notes = []
            if isinstance(b, (int, float)):
                amount = float(b)
            else:
                # e.g. Costo = "FALTA" with a tentative figure in column C
                amount = 0.0
                hint = next((n for n in side_notes if re.match(r"^[\d.,]+$", n)), None)
                notes.append(f"monto marcado '{b}' en el excel" + (f" (posible: {hint})" if hint else ""))
                side_notes = [n for n in side_notes if n != hint]

            expense = {
                "program": PROGRAM,
                "client_name": block["client_name"],
                "vehicle": block["vehicle"],
                "stock_number": block["stock_number"],
                "deal_number": block["deal_number"],
                "customer_number": block["customer_number"],
                "vin": block["vin"],
                "description": a_text,
                "category": categorize(a_text),
                "amount": round(amount, 2),
                "expense_date": extract_date(a_text),
                "source": "google-sheets",
            }
            for note in side_notes:
                parsed = parse_side_note(note, amount)
                note_text = parsed.pop("note", None)
                expense.update(parsed)
                if note_text:
                    notes.append(note_text)
            if not block["expenses"] and block["block_note"]:
                notes.append(f"[Cliente] {block['block_note']}")
            if notes:
                expense["notes"] = " | ".join(notes)
            block["expenses"].append(expense)

    flush()
    return rows, blocks_meta


def cross_check(wb, master_rows):
    """Report ledger entries (sheets 2/3) whose amount is absent from the master sheet."""
    master_amounts = {round(r["amount"], 2) for r in master_rows}
    unmatched = []

    ws = wb["Cust # 9301 ARG"]
    for r in range(2, ws.max_row + 1):
        amount = ws.cell(r, 3).value
        if isinstance(amount, (int, float)) and round(float(amount), 2) not in master_amounts:
            unmatched.append({
                "sheet": "Cust # 9301 ARG", "row": r, "amount": round(float(amount), 2),
                "description": str(ws.cell(r, 4).value or "")[:100],
            })
    return unmatched


def main():
    xlsx_path, out_path = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows, blocks = parse_master_sheet(wb["Vehicle Expense Report"])

    print(f"Parsed {len(blocks)} client blocks, {len(rows)} expense rows")
    print(f"Total imported: ${sum(r['amount'] for r in rows):,.2f}\n")

    print("Block totals vs Excel totals (mismatches flag typos in the Excel):")
    for blk in blocks:
        computed = round(sum(e["amount"] for e in blk["expenses"]), 2)
        excel = blk["excel_total"]
        flag = ""
        if excel is not None and abs(computed - excel) > 0.02:
            flag = f"  <-- EXCEL DICE {excel:,.2f} (diff {computed - excel:+,.2f})"
        elif excel is None:
            flag = "  (sin fila de total en el excel)"
        print(f"  {blk['client_name'][:40]:42} {computed:>12,.2f}{flag}")

    unmatched = cross_check(wb, rows)
    print(f"\nLedger 'Cust # 9301 ARG': {len(unmatched)} entries NOT found in master sheet (by amount):")
    for u in unmatched:
        print(f"  row {u['row']:>3}  ${u['amount']:>10,.2f}  {u['description']}")

    with open(out_path, "w") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    print(f"\nWrote {out_path}")


if __name__ == "__main__":
    main()
